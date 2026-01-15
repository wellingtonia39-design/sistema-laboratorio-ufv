import streamlit as st
import pandas as pd
import numpy as np
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from oauth2client.service_account import ServiceAccountCredentials
from fpdf import FPDF
import io
import os
import openpyxl
import json
from datetime import datetime

# --- CONFIGURAÇÃO ---
st.set_page_config(page_title="Sistema Controle UFV", layout="wide", page_icon="🌲")

# ✅ IDs CONFIGURADOS
ID_ARQUIVO_EXCEL = "1L0qTK6oy2axnCSlLadoyk9q5fExSnA6v"
ID_PASTA_RAIZ = "1nZtJjVZUVx65GtjnmpTn5Hw_eZOXwpIY"
ARQUIVO_CONFIG = "config_colunas_v54.json"

# --- DEFINIÇÃO DE COLUNAS PADRÃO (O QUE VEM MARCADO) ---
# Aqui definimos o "Kit Básico" para você não ter que marcar tudo do zero
COLS_PADRAO_MADEIRA = [
    "Selecionar", "Código UFV", "Data de entrada", "Nome do Cliente", "Aplicação", 
    "Grau", "Descrição Grau", # Para preencher o grau
    "Diâmetro 1 (mm)", "Diâmetro 2 (mm)", # Para digitar médias
    "Comprim. 1 (mm)", "Comprim. 2 (mm)",
    "Massa 1 (g)", "Massa 2 (g)",
    "Cromo (%)", "Cobre (%)", "Arsênio (%)", # Química
    "Retenção Total (Kg/m³)", "Observação" # Resultados
]

COLS_PADRAO_SOLUCAO = [
    "Código UFV", "Data de entrada", "Nome do Cliente",
    "Cromo (%)", "Cobre (%)", "Arsênio (%)",
    "Soma Concentração", "Balanço Total",
    "Grau do aspecto", "Descrição do aspecto"
]

# --- REGRAS ---
REGRAS_RETENCAO = {
    "Postes": 4.0, "Mourões": 6.5, "Dormentes": 6.5, "Cruzetas": 9.6, "Estacas": 6.5, "Madeira Serrada": 4.0
}
DESC_GRAU = {
    1: ("Profunda e regular", "Indica a penetração profunda e uniforme em toda a extensão do alburno."),
    2: ("Profunda e irregular", "Indica a penetração profunda, mas desuniforme em toda a extensão do alburno."),
    3: ("Parcial e regular", "Indica a penetração uniforme, mas não total pela extensão do alburno."),
    4: ("Parcial e irregular", "Indica a penetração desuniforme e não total pela extensão do alburno."),
    5: ("Sem Reação do Cromoazurol", "Sem Reação do Cromoazurol")
}
TXT_APROVADO = "Os resultados da análise química apresentaram uma retenção do produto de acordo com o padrão mínimo exigido pela norma ABNT NBR 16143"
TXT_REPROVADO = "Os resultados da análise química apresentaram uma retenção do produto inferior ao padrão mínimo exigido pela norma ABNT NBR 16143"

# --- PERSISTÊNCIA ---
def carregar_config():
    if os.path.exists(ARQUIVO_CONFIG):
        with open(ARQUIVO_CONFIG, "r") as f: return json.load(f)
    return {}

def salvar_config(config):
    with open(ARQUIVO_CONFIG, "w") as f: json.dump(config, f)

# --- DRIVE ---
def get_drive_service():
    scope = ["https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
    return build('drive', 'v3', credentials=creds)

def get_or_create_folder(service, folder_name, parent_id):
    query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and '{parent_id}' in parents and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
    items = results.get('files', [])
    if items: return items[0]['id']
    else:
        metadata = {'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [parent_id]}
        return service.files().create(body=metadata, fields='id', supportsAllDrives=True).execute().get('id')

def salvar_pdf_organizado(pdf_bytes, nome_arquivo, data_entrada_raw):
    try:
        if not ID_PASTA_RAIZ: st.error("⚠️ ID da pasta não configurado."); return
        service = get_drive_service()
        meses = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'}
        data_obj = datetime.now()
        if pd.isna(data_entrada_raw) or str(data_entrada_raw).strip() in ["", "NaT", "None"]: pass
        elif isinstance(data_entrada_raw, datetime): data_obj = data_entrada_raw
        else:
            try:
                v_str = str(data_entrada_raw).strip().split(" ")[0]
                for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]:
                    try: data_obj = datetime.strptime(v_str, fmt); break
                    except: continue
            except: pass
        ano_str = str(data_obj.year); mes_str = meses[data_obj.month]
        ano_id = get_or_create_folder(service, ano_str, ID_PASTA_RAIZ)
        mes_id = get_or_create_folder(service, mes_str, ano_id)
        nome_limpo = nome_arquivo.replace("/", "-").replace("\\", "-")
        media = MediaIoBaseUpload(io.BytesIO(pdf_bytes), mimetype='application/pdf', resumable=False)
        metadata = {'name': nome_limpo, 'parents': [mes_id]}
        service.files().create(body=metadata, media_body=media, fields='id', supportsAllDrives=True).execute()
        st.balloons(); st.toast(f"Salvo: {ano_str}/{mes_str}", icon="✅"); st.success(f"Arquivo **{nome_limpo}** salvo em: **{ano_str} > {mes_str}**")
    except Exception as e: st.error(f"Erro ao salvar: {e}")

# --- MATEMÁTICA ---
def to_float(v):
    try: return 0.0 if (pd.isna(v) or str(v).strip()=="") else float(str(v).replace(",", "."))
    except: return 0.0

def aplicar_formulas_excel(df):
    # Percorre linha a linha para recalcular TUDO baseado nas entradas
    for i, row in df.iterrows():
        try:
            # 1. GRAU (Preenchimento de Texto)
            if 'Grau' in df.columns:
                grau = to_float(row.get('Grau'))
                if grau > 0 and int(grau) in DESC_GRAU:
                    d_curta, d_longa = DESC_GRAU[int(grau)]
                    df.at[i, 'Descrição Grau'] = d_curta
                    df.at[i, 'Descrição Penetração'] = d_longa

            # 2. MÉDIAS FÍSICAS (Se existirem as colunas de entrada)
            if 'Diâmetro 1 (mm)' in df.columns:
                # Diâmetro
                d_list = [to_float(row.get(f'Diâmetro {x} (mm)')) for x in range(1,6)]
                diams = [d for d in d_list if d > 0]
                diam_medio_cm = (sum(diams)/len(diams))/10.0 if diams else 0
                df.at[i, 'Diâmetro médio (cm)'] = round(diam_medio_cm, 2)

                # Comprimento
                c_list = [to_float(row.get(f'Comprim. {x} (mm)')) for x in range(1,6)]
                comps = [c for c in c_list if c > 0]
                comp_medio_cm = (sum(comps)/len(comps))/10.0 if comps else 0
                df.at[i, 'Comprim. Médio (cm)'] = round(comp_medio_cm, 2)

                # Massa
                m_list = [to_float(row.get(f'Massa {x} (g)')) for x in range(1,6)]
                massas = [m for m in m_list if m > 0]
                massa_media = sum(massas)/len(massas) if massas else 0
                df.at[i, 'Massa média (g)'] = round(massa_media, 2)

                # Volume e Densidade
                dens_kg_m3 = 0
                if diam_medio_cm > 0 and comp_medio_cm > 0:
                    vol = 3.14159 * ((diam_medio_cm/2)**2) * comp_medio_cm
                    df.at[i, 'Volume (cm³)'] = round(vol, 2)
                    if massa_media > 0:
                        dens_g_cm3 = massa_media / vol
                        dens_kg_m3 = dens_g_cm3 * 1000
                        df.at[i, 'Densidade (g/cm³)'] = round(dens_g_cm3, 3)
                        df.at[i, 'Densidade (Kg/m³)'] = round(dens_kg_m3, 2)
                
                # 3. QUÍMICA
                cr_pct = to_float(row.get('Cromo (%)'))
                cu_pct = to_float(row.get('Cobre (%)'))
                as_pct = to_float(row.get('Arsênio (%)'))
                soma_conc = cr_pct + cu_pct + as_pct
                df.at[i, 'Soma Concentração'] = round(soma_conc, 2)

                if soma_conc > 0:
                    df.at[i, 'Balanço Cromo %'] = round((cr_pct/soma_conc)*100, 2)
                    df.at[i, 'Balanço Cobre %'] = round((cu_pct/soma_conc)*100, 2)
                    df.at[i, 'Balanço Arsênio %'] = round((as_pct/soma_conc)*100, 2)
                    df.at[i, 'Balanço Total'] = 100.00
                
                ret_cr = (cr_pct/100)*dens_kg_m3
                ret_cu = (cu_pct/100)*dens_kg_m3
                ret_as = (as_pct/100)*dens_kg_m3
                
                df.at[i, 'Retenção Cromo (Kg/m³)'] = round(ret_cr, 2)
                df.at[i, 'Retenção Cobre (Kg/m³)'] = round(ret_cu, 2)
                df.at[i, 'Retenção Arsênio (Kg/m³)'] = round(ret_as, 2)
                
                ret_total = ret_cr + ret_cu + ret_as
                df.at[i, 'Retenção Total (Kg/m³)'] = round(ret_total, 2)

                # 4. APROVAÇÃO E REGRAS
                aplicacao = str(row.get('Aplicação', '')).strip()
                ret_esp = 0.0
                for k, v in REGRAS_RETENCAO.items():
                    if k.lower() in aplicacao.lower(): ret_esp = v; break
                
                if ret_esp > 0:
                    df.at[i, 'Retenção'] = ret_esp
                    df.at[i, 'Retenção Esp.'] = ret_esp
                    df.at[i, 'Observação'] = TXT_APROVADO if ret_total >= ret_esp else TXT_REPROVADO

        except: continue
    return df

@st.cache_data(ttl=60)
def carregar_excel_drive(aba_nome):
    try:
        service = get_drive_service()
        request = service.files().get_media(fileId=ID_ARQUIVO_EXCEL)
        df = pd.read_excel(io.BytesIO(request.execute()), sheet_name=aba_nome)
        df.columns = df.columns.str.strip()
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Filtros de Limpeza (mantidos da V52)
        if aba_nome == "Madeira Tratada":
            cols_proibidas = ['pH da solução', 'Densidade  solução (g/cm³)', 'Temperatura', 'Concentração pela tabela']
            df = df.drop(columns=[c for c in cols_proibidas if c in df.columns], errors='ignore')
        elif aba_nome == "Solução Preservativa":
            cols_proibidas = ['Diâmetro 1 (mm)', 'Massa 1 (g)', 'Retenção', 'Retenção Esp.'] # Simplificado
            df = df.drop(columns=[c for c in cols_proibidas if c in df.columns], errors='ignore')
            
        return df
    except Exception as e: st.error(f"Erro Excel: {e}"); return pd.DataFrame()

def salvar_excel_drive(df_to_save, aba_nome):
    try:
        # 1. Aplica Matemática
        df_final = aplicar_formulas_excel(df_to_save)
        
        # 2. Salva Preservando Formatação
        service = get_drive_service()
        request = service.files().get_media(fileId=ID_ARQUIVO_EXCEL)
        arquivo_original = io.BytesIO(request.execute())
        wb = openpyxl.load_workbook(arquivo_original)
        if aba_nome not in wb.sheetnames: st.error("Aba não encontrada"); return
        ws = wb[aba_nome]
        
        col_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1) if cell.value}
        col_id_idx = col_map.get("Código UFV")
        
        excel_rows = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[col_id_idx - 1]: excel_rows[str(row[col_id_idx - 1]).strip()] = row_idx
            
        for index, row_df in df_final.iterrows():
            codigo = str(row_df.get('Código UFV','')).strip()
            linha = excel_rows.get(codigo)
            if linha:
                for col, val in row_df.items():
                    c_idx = col_map.get(col)
                    if c_idx: ws.cell(row=linha, column=c_idx, value=val)
        
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        media = MediaIoBaseUpload(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        service.files().update(fileId=ID_ARQUIVO_EXCEL, media_body=media, supportsAllDrives=True).execute()
        st.toast("Calculado e Salvo!", icon="✅"); st.cache_data.clear()
        
    except Exception as e: st.error(f"Erro Salvar: {e}")

# --- PDF E HELPERS ---
def clean_text(text): return str(text).encode('latin-1', 'replace').decode('latin-1') if not pd.isna(text) else ""
def fmt_num(v): 
    try: return "{:,.2f}".format(float(str(v).replace(",", "."))).replace(",", "X").replace(".", ",").replace("X", ".")
    except: return str(v)
def fmt_date(v):
    if pd.isna(v) or v is None or str(v).strip() in ["", "NaT", "None"]: return "-"
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    s = str(v).strip().split(" ")[0]
    for f in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
        try: return datetime.strptime(s, f).strftime("%d/%m/%Y")
        except: continue
    return s
def get_val(d, keys):
    dn = {k.strip().lower(): v for k, v in d.items()}
    for k in keys:
        if k.strip().lower() in dn:
            val = dn[k.strip().lower()]
            if not pd.isna(val) and str(val).strip() not in ["", "NaT"]: return val
    return ""

class RPDF(FPDF):
    def header(self):
        if os.path.exists("logo_ufv.png"): self.image("logo_ufv.png", 10, 8, 25)
        if os.path.exists("logo_montana.png"): self.image("logo_montana.png", 155, 8, 45) 
        self.set_y(12); self.set_font('Arial','B',14); self.cell(0,10,clean_text('Relatório de Ensaio'),0,1,'C')
    def footer(self): self.set_y(-15); self.set_font('Arial','I',6); self.cell(0,10,clean_text(f'Página {self.page_no()}'),0,0,'C')
    def field(self, label, valor, x, y, w, h=6, align='L', multi=False, bold_value=False):
        self.set_xy(x, y); self.set_font('Arial', 'B', 8); self.cell(w, 3, clean_text(label), 0, 0, 'L')
        self.set_xy(x, y+3)
        if bold_value: self.set_font('Arial', 'B', 8)
        else: self.set_font('Arial', '', 8)
        if multi: self.rect(x, y+3, w, h); self.multi_cell(w, 4, clean_text(valor), 0, align)
        else: self.cell(w, h, clean_text(valor), 1, 0, align)
    def draw_chem_label(self, tipo):
        x_start, y_start = self.get_x(), self.get_y(); self.set_font('Arial', '', 8)
        def write_part(txt, size=8, offset_y=0):
            self.set_font('Arial', '', size); w = self.get_string_width(txt); curr_x = self.get_x()
            self.set_xy(curr_x, y_start + offset_y); self.cell(w, 6, clean_text(txt), 0, 0); self.set_xy(curr_x + w, y_start)
        if tipo == "Cr": write_part("Teor de CrO"); write_part("3", size=5, offset_y=1.5); write_part(" (Cromo)")
        elif tipo == "Cu": write_part("Teor de CuO (Cobre)")
        elif tipo == "As": write_part("Teor de As"); write_part("2", size=5, offset_y=1.5); write_part("O"); write_part("5", size=5, offset_y=1.5); write_part(" (Arsênio)")
        self.set_xy(x_start, y_start); self.cell(40, 6, "", 1, 0)

def gerar_pdf(d):
    pdf = RPDF(); pdf.add_page(); pdf.set_auto_page_break(auto=True, margin=15)
    y = 30
    pdf.field("Data de Entrada", fmt_date(get_val(d, ["Data de entrada", "Entrada"])), 10, y, 40, align='C')
    pdf.field("Número ID", clean_text(get_val(d, ["Código UFV", "ID"])), 150, y-5, 50, align='C')
    pdf.field("Data de Emissão", fmt_date(get_val(d, ["Data de Registro", "Fim da análise"])), 150, y+8, 50, align='C')
    y += 20; pdf.set_y(y); pdf.set_font('Arial', 'B', 9); pdf.cell(0, 5, clean_text("DADOS DO CLIENTE"), 0, 1, 'L')
    y += 6; pdf.field("Cliente", get_val(d, ["Nome do Cliente"]), 10, y, 190)
    y += 11; pdf.field("Cidade/UF", f"{get_val(d,['Cidade'])}/{get_val(d,['Estado'])}", 10, y, 90)
    pdf.field("E-mail", get_val(d, ["E-mail"]), 105, y, 95)
    y += 15; pdf.set_y(y); pdf.set_font('Arial', 'B', 9); pdf.cell(0, 5, clean_text("IDENTIFICAÇÃO DA AMOSTRA"), 0, 1, 'L')
    y += 6; pdf.field("Ref. Cliente", get_val(d, ["Indentificação de Amostra"]), 10, y, 190)
    y += 11; pdf.field("Madeira", get_val(d, ["Madeira"]), 10, y, 90)
    pdf.field("Produto", get_val(d, ["Produto"]), 105, y, 95)
    y += 11; pdf.field("Aplicação", get_val(d, ["Aplicação"]), 10, y, 60)
    pdf.field("Norma ABNT", get_val(d, ["Norma"]), 75, y, 60)
    ret_esp = get_val(d, ["Retenção", "Retenção Esp."])
    pdf.field("Retenção Esp.", fmt_num(ret_esp), 140, y, 60, align='C')
    y += 20; pdf.set_y(y); pdf.set_font('Arial', 'B', 9); pdf.cell(190, 6, clean_text("RESULTADOS DE RETENÇÃO"), 1, 1, 'C')
    pdf.set_font('Arial', 'B', 7); x=10; cy=pdf.get_y()
    pdf.cell(40, 10, clean_text("Ingredientes ativos"), 1, 0, 'C'); pdf.cell(30, 10, clean_text("Resultado (kg/m3)"), 1, 0, 'C'); pdf.cell(80, 5, clean_text("Balanceamento químico"), 1, 0, 'C')
    pdf.set_xy(x+150, cy); pdf.cell(40, 10, clean_text("Método"), 1, 0, 'C'); pdf.set_xy(x+70, cy+5); pdf.cell(30, 5, clean_text("Resultados (%)"), 1, 0, 'C'); pdf.cell(50, 5, clean_text("Padrões"), 1, 0, 'C')
    pdf.set_xy(x, cy+10); y_dados_inicio = cy+10
    kg_cr=fmt_num(get_val(d,["Retenção Cromo (Kg/m³)","Retenção Cromo"])); kg_cu=fmt_num(get_val(d,["Retenção Cobre (Kg/m³)","Retenção Cobre"])); kg_as=fmt_num(get_val(d,["Retenção Arsênio (Kg/m³)","Retenção Arsênio"]))
    pc_cr=fmt_num(get_val(d,["Balanço Cromo %","Balanço Cromo"])); pc_cu=fmt_num(get_val(d,["Balanço Cobre %","Balanço Cobre"])); pc_as=fmt_num(get_val(d,["Balanço Arsênio %","Balanço Arsênio"]))
    pdf.set_font('Arial', '', 8)
    def row_data_custom(tipo, k, p, mn, mx):
        pdf.draw_chem_label(tipo); pdf.cell(30, 6, k, 1, 0, 'C'); pdf.cell(30, 6, p, 1, 0, 'C'); pdf.cell(25, 6, mn, 1, 0, 'C'); pdf.cell(25, 6, mx, 1, 0, 'C'); pdf.set_x(pdf.get_x() + 40); pdf.ln(6)
    pdf.set_xy(160, y_dados_inicio); pdf.cell(40, 18, clean_text("Metodo UFV 01"), 1, 0, 'C'); pdf.set_xy(10, y_dados_inicio)
    row_data_custom("Cr", kg_cr, pc_cr, "41,8", "53,2"); row_data_custom("Cu", kg_cu, pc_cu, "15,2", "22,8"); row_data_custom("As", kg_as, pc_as, "27,3", "40,7")
    try: tot_kg = float(kg_cr.replace(",",".")) + float(kg_cu.replace(",",".")) + float(kg_as.replace(",","."))
    except: tot_kg = 0
    try: soma_pct = float(pc_cr.replace(",",".")) + float(pc_cu.replace(",",".")) + float(pc_as.replace(",","."))
    except: soma_pct = 100.00
    pdf.set_font('Arial', 'B', 8); pdf.cell(40, 6, clean_text("RETENÇÃO TOTAL"), 1, 0, 'L'); pdf.cell(30, 6, fmt_num(tot_kg), 1, 0, 'C'); pdf.cell(30, 6, fmt_num(soma_pct), 1, 0, 'C'); pdf.cell(90, 6, clean_text("Nota: Resultados restritos as amostras"), 1, 1, 'C')
    y = pdf.get_y() + 5; pdf.set_y(y); pdf.set_font('Arial', 'B', 9); pdf.cell(190, 6, clean_text("RESULTADOS DE PENETRAÇÃO"), 0, 1, 'C'); y += 7
    tipo_correto = get_val(d, ["Descrição Grau", "Descrição do Grau", "Grau Descricao"])
    pdf.field("Grau", get_val(d, ["Grau"]), 10, y, 30, align='C'); pdf.field("Tipo", tipo_correto, 45, y, 50, align='C')
    pdf.set_xy(100, y); pdf.set_font('Arial', 'B', 8); pdf.cell(90, 3, clean_text("Descrição"), 0, 0, 'L'); pdf.set_xy(100, y+3); pdf.set_font('Arial', '', 8); pdf.rect(100, y+3, 100, 12); pdf.multi_cell(100, 4, clean_text(get_val(d, ["Descrição Penetração"])), 0, 'L')
    y += 20; obs = get_val(d, ["Observação", "Obs"])
    if obs: pdf.set_y(y); pdf.field("Observações", obs, 10, y, 190, 12, 'L', multi=True, bold_value=True)
    pdf.set_y(-35); pdf.set_font('Arial', '', 9); pdf.cell(0, 5, clean_text("Dr. Vinicius Resende de Castro - Supervisor do laboratório"), 0, 1, 'C')
    return pdf.output(dest='S').encode('latin-1')

# --- MAIN ---
def main():
    if 'logado' not in st.session_state: st.session_state['logado']=False
    if not st.session_state['logado']:
        c1,c2,c3=st.columns([1,2,1])
        with c2:
            st.title("🔐 Login"); u=st.text_input("User"); p=st.text_input("Pass",type="password")
            if st.button("Entrar",type="primary"):
                if (u=="admin" and p=="admin") or (u=="montana" and p=="montana"): st.session_state.update({'logado':True,'tipo':u.capitalize(),'user':u}); st.rerun()
                else: st.error("Erro")
        return
    st.sidebar.info(f"👤 {st.session_state['user']}"); 
    if st.sidebar.button("Sair"): st.session_state['logado']=False; st.rerun()
    st.title("🌲 Sistema Controle UFV")
    menu=st.sidebar.radio("Menu",["Madeira Tratada","Solução"])
    
    # --- CONFIGURAÇÃO DE COLUNAS ---
    config = carregar_config()

    if menu=="Madeira Tratada":
        df=carregar_excel_drive("Madeira Tratada")
        if not df.empty:
            if "Selecionar" not in df.columns: df.insert(0,"Selecionar",False)
            
            # --- SELETOR DE COLUNAS ---
            cols_disponiveis = [c for c in df.columns if c not in ["Selecionar", "Código UFV"]]
            
            # Tenta pegar config salva ou usa o Padrão Limpo
            padrao = [c for c in COLS_PADRAO_MADEIRA if c in cols_disponiveis]
            escolha_usuario = config.get("Madeira", padrao)
            
            with st.expander("⚙️ Personalizar Colunas (Adicionar/Remover)"):
                cols_visiveis = st.multiselect("Marque as colunas que deseja ver:", cols_disponiveis, default=escolha_usuario)
                if st.button("💾 Salvar Preferência"):
                    config["Madeira"] = cols_visiveis
                    salvar_config(config)
                    st.success("Preferência Salva!")
                    st.rerun()

            # Lista final de colunas (Obrigatórias + Escolhidas)
            cols_finais = ["Selecionar", "Código UFV"] + cols_visiveis
            # Filtra colunas que realmente existem no DF para evitar erro
            cols_finais = [c for c in cols_finais if c in df.columns]

            st.markdown("### 🔎 Buscar/Editar Amostra")
            col_busca, col_info = st.columns([1, 3])
            with col_busca: numero_busca = st.text_input("Digite o número (ex: 620)", placeholder="Busque para Editar...")
            
            if numero_busca:
                termo = f"UFV-M-{numero_busca}"
                df_filtrado = df[df['Código UFV'].astype(str).str.contains(termo, case=False, na=False)]
                if df_filtrado.empty: df_filtrado = df[df['Código UFV'].astype(str).str.contains(numero_busca, case=False, na=False)]
                
                with col_info: st.info(f"Encontrados: {len(df_filtrado)}. Edite e clique em CALCULAR.")
                
                df_view = st.data_editor(df_filtrado[cols_finais], num_rows="dynamic", use_container_width=True, key="tabela_filtrada")
                
                if st.session_state['user'] in ["admin", "Lpm"]:
                    if st.button("🧮 CALCULAR E SALVAR (Mesclar)", type="primary"):
                        df.update(df_view)
                        salvar_excel_drive(df, "Madeira Tratada")
                        st.success("Atualizado!")
                        st.rerun()
            else:
                with col_info: st.info("Mostrando tabela completa.")
                df_view = st.data_editor(df[cols_finais], num_rows="dynamic", use_container_width=True, key="tabela_completa")
                
                if st.session_state['user'] in ["admin", "Lpm"]:
                    if st.button("🧮 CALCULAR E SALVAR TUDO", type="primary"): 
                        df.update(df_view)
                        salvar_excel_drive(df, "Madeira Tratada")
                        st.rerun()

            # PDF
            if numero_busca:
                # Recarrega a linha do DF original para pegar colunas ocultas
                current_df = df[df['Código UFV'].isin(df_filtrado['Código UFV'])]
                sel_codes = df_view[df_view['Selecionar']==True]['Código UFV'].tolist()
                sel = current_df[current_df['Código UFV'].isin(sel_codes)]
            else:
                sel_codes = df_view[df_view['Selecionar']==True]['Código UFV'].tolist()
                sel = df[df['Código UFV'].isin(sel_codes)]

            st.divider()
            if not sel.empty:
                st.subheader("📄 Gerar Relatório")
                try:
                    l=sel.iloc[0].to_dict()
                    pdf_bytes=gerar_pdf(l)
                    nome_arquivo = f"{l.get('Código UFV','Relatorio')}.pdf"
                    c_down, c_cloud = st.columns(2)
                    with c_down: st.download_button("⬇️ BAIXAR PDF (PC)", pdf_bytes, nome_arquivo, "application/pdf", type="primary")
                    with c_cloud:
                        if st.button("☁️ SALVAR NO DRIVE COMPARTILHADO"): salvar_pdf_organizado(pdf_bytes, nome_arquivo, get_val(l,["Data de entrada"]))
                except Exception as e: st.error(f"Erro na geração: {e}")
            else: 
                if numero_busca and df_filtrado.empty: st.warning("Nenhum resultado.")

    elif menu=="Solução":
        df=carregar_excel_drive("Solução Preservativa")
        if not df.empty:
            cols_disponiveis = [c for c in df.columns if c not in ["Código UFV"]]
            padrao = [c for c in COLS_PADRAO_SOLUCAO if c in cols_disponiveis]
            escolha_usuario = config.get("Solução", padrao)
            
            with st.expander("⚙️ Personalizar Colunas"):
                cols_visiveis = st.multiselect("Marque as colunas que deseja ver:", cols_disponiveis, default=escolha_usuario)
                if st.button("💾 Salvar Preferência Solução"):
                    config["Solução"] = cols_visiveis
                    salvar_config(config)
                    st.rerun()
            
            cols_finais = ["Código UFV"] + cols_visiveis
            cols_finais = [c for c in cols_finais if c in df.columns]
            st.data_editor(df[cols_finais], use_container_width=True)

if __name__ == "__main__":
    main()
